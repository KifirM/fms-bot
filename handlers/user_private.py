import json
import os

from aiogram import types, Router, F
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, Command, or_f
from aiogram.utils.formatting import as_list, as_marked_section, Bold
from database.schedule_conversion import all_day

from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

from filters.chat_types import ChatTypeFilter
from keyboards import reply
from database.requests import del_task, set_task

from openpyxl import load_workbook
from database.schedule_conversion import get_file

from aiogram.types import Message, FSInputFile

from datetime import date, timedelta,datetime,timezone
import pytz

user_private_router = Router()
user_private_router.message.filter(ChatTypeFilter(['private']))

liters = ['Μ', 'Ξ', 'Ο', 'Π', 'Ρ', 'Σ', 'Τ', 'Φ', 'Χ', 'Ψ', 'Β', 'Ζ', 'Γ', 'Ι', 'Δ', 'Η', 'Θ', 'Ε', 'К', 'Λ']
groups = ['B', 'A']
classes = ['10', '11']


async def download():
    kras_timezone = pytz.timezone('Asia/Krasnoyarsk')
    current_hour_kras = datetime.now(kras_timezone).hour
    with open(f'date.txt', 'r', encoding="utf-8") as f:
        datt = f.read()
        datt = datetime.strptime(datt, '%Y-%m-%d').date()
    dt_to = date.today()

    if datt > dt_to:
        return 0
    elif datt == dt_to:
        if current_hour_kras >= 16:
            dt_to = date.today() + timedelta(days=1)
            get_file(dt_to)
            get_time_tab()
            return 1
        else:
            return 0
    elif datt < dt_to:
        if current_hour_kras >= 16:
            dt_to = date.today() + timedelta(days=1)
            get_file(dt_to)
            get_time_tab()
            return 1
        else:
            get_file(dt_to)
            get_time_tab()
            return 1

from aiogram.utils.keyboard import ReplyKeyboardBuilder, InlineKeyboardBuilder, InlineKeyboardButton

def get_time_tab():
    workbook = load_workbook('Sample.xlsx')
    for el in workbook.sheetnames:
        if '11' in el:
            all_day('11')
        if '10' in el:
            all_day('10')

def class_variants():
    with open(f'data_10.json', 'r', encoding="utf-8") as f:
        file = json.load(f)
        button_list = file['classes']['10']
    with open(f'data_11.json', 'r', encoding="utf-8") as f:
        file = json.load(f)
        button_list = button_list + file['classes']['11']
    return button_list


@user_private_router.message(CommandStart())
async def start_cmd(message: types.Message):
    await message.answer(
        'Привет! Я — ФМШ бот, твой персональный помощник 😉\nC моей помощью ты сможешь создавать заметки, которые всегда будут под рукой и смотреть расписание в удобном формате',
        reply_markup=reply.start_kb.as_markup(
            resize_keyboard=True,
            input_field_placeholder='Что тебя интересует?'))


@user_private_router.message(F.text.lower() == 'меню 📂')
async def menu_cmd(message: types.Message):
    await message.answer('Выбери действие:',
                         reply_markup=reply.menu_kb.as_markup(resize_keyboard=True))


@user_private_router.message(F.text == 'Вернуться в начало 🏠')
async def start_cmd(message: types.Message):
    await message.answer('Ты вернулся в начало', reply_markup=reply.start_kb.as_markup(resize_keyboard=True))


@user_private_router.message(F.text.lower() == 'заметки 📝')
async def notes_cmd(message: types.Message):
    await message.answer('Нажми на выполненную задачу, чтобы удалить или напиши новую',
                         reply_markup=await reply.tasks(message.from_user.id))


# @user_private_router.callback_query(F.data.startswith('task_'))
# async def delete_task(callback: types.CallbackQuery):
#     await callback.answer('Задача удалена')
#     await del_task(callback.data.split('_')[1])
#     await callback.message.delete()
#     await callback.message.answer('Нажми на выполненную задачу, чтобы удалить или напиши новую',
#                                   reply_markup=await reply.tasks(callback.from_user.id))
#     await callback.answer('Задача выполнена!')

# @user_private_router.message(or_f(F.text.lower() != 'контакты 📞'), (F.text.lower() != '/restart'), (F.text.lower() != 'помощь ❓'), (F.text.lower() != 'расписание 📆'), (F.text.lower() != '⬅️ назад'))
# async def add_task(message: types.Message):
#     if len(message.text) > 200:
#         await message.answer('Задача слишком длинная')
#         return
#     await set_task(message.from_user.id, message.text)
#     await message.answer('Задача успешно добавлена\nНажми на выполненную задачу, чтобы удалить или напиши новую',
#                          reply_markup=await reply.tasks(message.from_user.id))


# @user_private_router.message(F.text.lower() == 'расписание 📆')
# async def schedule_cmd(message: types.Message):
#     await set_task(message.from_user.id, message.text)
#     await message.answer('Выбери класс',
#                          reply_markup=await reply.tasks(message.from_user.id))

@user_private_router.message(F.text.lower() == 'расписание 📆')
async def class_cmd(message: types.Message):
    await download() 
    await message.answer('выбери класс', reply_markup=reply.clases_kb.as_markup(resize_keyboard=True))


@user_private_router.message(F.text == 'Выбрать другую литеру 🔠')
async def back_cmd(message: types.Message, state: FSMContext):    
    if await download():
        await state.clear()
    data = await state.get_data()
    if len(data.keys()) == 2 or len(data.keys()) == 3:
        if data['user_class']:
            await send_button(message, state)
        else:
            await state.clear()
            await message.answer('выбери класс', reply_markup=reply.clases_kb.as_markup(resize_keyboard=True))
    else:
        await state.clear()
        await message.answer('выбери класс', reply_markup=reply.clases_kb.as_markup(resize_keyboard=True))


@user_private_router.message(F.text == 'Вернуться в меню 📂🔙')
async def back_cmd(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer('Выбери действие:',
                         reply_markup=reply.menu_kb.as_markup(resize_keyboard=True))

@user_private_router.message(F.text == '⚙️')
async def back_cmd(message: types.Message, state: FSMContext):
    if message.from_user.id == 5480167477 or message.from_user.id == 1550008797:
        await message.answer('Выбери действие:',
                             reply_markup=reply.admin_kb.as_markup(resize_keyboard=True))
    else:
        await message.answer('Доступ закрыт.')


@user_private_router.message(F.text.lower().in_([classs.lower() for classs in classes]))
async def send_button(message: types.Message, state: FSMContext):
    clas = message.text
    if await download():
        await state.set_state()
    if clas == '10' or clas == '11':
        await state.update_data(user_class=clas)
    else:
        data = await state.get_data()
        clas = data['user_class']
    with open(f'data_{clas}.json', 'r', encoding="utf-8") as f:
        file = json.load(f)
        button_list = file['classes'][clas]
    liter = ReplyKeyboardBuilder()
    for button_text in button_list:
        liter.button(text=button_text)
    liter.button(text='🔀 Выбрать другой класс 🏫'),
    liter.button(text='Вернуться в меню 📂🔙'),
    liter.button(text='Вернуться в начало 🏠')
    liter.adjust(5)
    await message.answer("Выберите литеру:", reply_markup=liter.as_markup(resize_keyboard=True))


@user_private_router.message(F.text == '🔀 Выбрать другой класс 🏫')
async def back_cmd(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer('выбери класс', reply_markup=reply.clases_kb.as_markup(resize_keyboard=True))


@user_private_router.message(lambda message: message.text.lower() in [var.lower() for var in class_variants()])
async def back_cmd(message: types.Message, state: FSMContext):
    data = await state.get_data()
    if await download():
        await state.set_state()
    if len(data.keys()) >= 1:
        await state.update_data(user_litera=message.text)
        await message.answer('выбери группу', reply_markup=reply.group_kb.as_markup(resize_keyboard=True))
    else:
        await state.clear()
        await message.answer('выбери класс', reply_markup=reply.clases_kb.as_markup(resize_keyboard=True))


@user_private_router.message(F.text.lower().in_([group.lower() for group in groups]))
async def group_A_cmd(message: types.Message, state: FSMContext):
    await state.update_data(user_group=message.text)
    data = await state.get_data()
    if await  download():
        await state.set_state()
    if len(data.keys()) == 3:
        with open(f'data_{data['user_class']}.json', 'r', encoding="utf-8") as f:
            file = json.load(f)
        print(data['user_litera'])
        await message.answer(
            f'Рассписание на {file['date']}\n{data['user_litera']} {data['user_group']}{file[data['user_group']][data['user_litera']]}')

    else:
        await state.clear()
        await message.answer('выбери класс', reply_markup=reply.clases_kb.as_markup(resize_keyboard=True))


@user_private_router.message(F.text == '🔄')
async def reload_data_cmd(message: types.Message, state: FSMContext):
    if message.from_user.id == 5480167477 or message.from_user.id == 1550008797:
        await message.answer('Сохраняю рассписание.')
        get_file(date.today())
        get_time_tab()
        await message.answer('Рассписание обновлено.')



#
# @user_private_router.callback_query(F.data == '10')
# async def tenth_classes(message: types.Message):
#     await message.answer(text='10 класс')
#
# @user_private_router.callback_query(F.data == '11')
# async def tenth_classes(message: types.Message):
#     await message.answer(text= '11 класс',reply_markup=reply.liter_kb_11())


@user_private_router.message(F.text.lower() == 'помощь ❓')
async def help_cmd(message: types.Message):
    text = '<b>Руководство пользователя</b>\n🟢 Основной функционал бота будет доступен после нажатия на кнопку "Меню".\n🟢 Для связи с разработчиком выбери кнопку "Контакты".\n🟢 Оценить бота ты сможешь, пройдя короткий опрос по ссылке: https://forms.gle/DKfZa3WKFjWcVGKVA'
    await message.answer(text)


@user_private_router.message(F.text.lower() == 'контакты 📞')
async def contacts_cmd(message: types.Message):
    await message.answer('<b>Контакты:</b>\nПо вопросам и предложениям — @k1f1r1k @ag_st_d', parse_mode='HTML')

@user_private_router.message(F.text == '📒')
async def cmd_start(message: Message, state: FSMContext):
    if message.from_user.id == 5480167477 or message.from_user.id == 1550008797:
        db_file = FSInputFile('db.sqlite3')
        await message.answer_document(db_file)

from openpyxl import load_workbook
import json
import requests
from bs4 import BeautifulSoup



def get_file():
    from datetime import date
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.164 Safari/537.36'
    session = requests.Session()
    session.headers.update({'User-Agent': user_agent})

    url = "https://fms.eljur.ru/authorize"
    session.headers.update({'Referer': url})

    data = {
        "username": '89504075585',
        "password": '89504075585'
    }
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36"}

    # авторизация
    ss = session.post("https://fms.eljur.ru/ajaxauthorize", data=data, headers=headers)
    date = str(date.today()).split('-')
    # date = f'{date[2]}.{date[1]}'
    date = f'20.12'

    json_string = session.get('https://fms.eljur.ru/journal-board-action')
    bs = BeautifulSoup(json_string.text, "html.parser")
    d = [el['href'] for el in bs.find_all("a", {"title": f"{date}.xlsx"})]
    if not d:
        print('рассписание отсутствует')
        exit()

    response = session.get(d[0])
    file_Path = 'Sample.xlsx'

    if response.status_code == 200:
        with open(file_Path, 'wb') as file:
            file.write(response.content)
        print('File downloaded successfully')
    else:
        print('Failed to download file')

    session.close()

def read_excel_with_merged_cells(excel_file,classs):


    workbook = load_workbook(excel_file)
    for el in workbook.sheetnames:
        if classs in el:
            classs = el
            break
    workbook.active = workbook[classs]
    sheet = workbook.active
    table_data = []
    max_row = sheet.max_row
    max_col = sheet.max_column

    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value

            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row = merged_range.min_col, merged_range.min_row
                max_col, max_row = merged_range.max_col, merged_range.max_row

                if min_row <= row <= max_row and min_col <= col <= max_col:
                    top_left_cell = sheet.cell(row=min_row, column=min_col)
                    cell_value = top_left_cell.value
                    break

            row_data.append(cell_value)
        table_data.append(row_data)

    return table_data

def all_day():
    get_file()
    sort_table_10 = read_excel_with_merged_cells("Sample.xlsx",'10')
    sort_table_10 = sort_table_10[1:]
    for i in range(len(sort_table_10)):
        sort_table_10[i].pop(0)


    sort_table_11 = read_excel_with_merged_cells("Sample.xlsx",'11')
    sort_table_11 = sort_table_11[1:]
    for i in range(len(sort_table_11)):
        sort_table_11[i] = sort_table_11[i][3:]

    sort_table_all = []

    for i in range(len(sort_table_11)):
        sort_table_all.append(sort_table_10[i] + sort_table_11[i])
    print(sort_table_all)




    classes_A = {}
    classes_B = {}

    for el in sort_table_all[0]:
        if '1' in el and el not in classes_A and el not in classes_A:
            classes_A[el] = None
            classes_B[el] = None



    for cl in list(classes_A.keys()):
        rasp_A = []
        rasp_B = []

        for el in sort_table_all[2:]:
            rasp_A.append(f'{el[0]} {el[1]} {el[sort_table_all[0].index(cl)]}')
            rasp_B.append(f'{el[0]} {el[1]} {el[sort_table_all[0].index(cl)+1]}')
        classes_A[cl] = rasp_A
        classes_B[cl] = rasp_B
        data = {'A' : classes_A , 'B' : classes_B}

    with open('data.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

all_day()


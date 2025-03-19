from openpyxl import load_workbook
import json

from bs4 import BeautifulSoup

from datetime import date, timedelta,datetime,timezone
import pytz


# kras_timezone = pytz.timezone('Asia/Krasnoyarsk')
# current_time_kras = datetime.now(kras_timezone)
# print(current_time_kras.hour)
def create_date_list_with_step(start_date, end_date, step_days):
  date_list = []
  current_date = start_date
  while current_date <= end_date:
      date_list.append(current_date)
      current_date += timedelta(days=step_days)
  return reversed(date_list)




def get_file(dt_to):
    import pytz
    import requests
    from datetime import date
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.164 Safari/537.36'
    session = requests.Session()
    session.headers.update({'User-Agent': user_agent})

    url = "https://fms.eljur.ru/authorize"
    session.headers.update({'Referer': url})

    data = {
        "username": '89504075585',
        "password": '89504075585'
    }
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36"}

    # авторизация
    ss = session.post("https://fms.eljur.ru/ajaxauthorize", data=data, headers=headers)
    kras_timezone = pytz.timezone('Asia/Krasnoyarsk')
    current_hour_kras = datetime.now(kras_timezone).hour
    dat = str(dt_to).split('-')
    datee = f'{dat[2]}.{dat[1]}'
    data = None

    for dt in create_date_list_with_step(start_date=date(int(dat[0])-1, 9, 1), end_date=dt_to, step_days=1):

        print(dt)
        json_string = session.get('https://fms.eljur.ru/journal-board-action')
        bs = BeautifulSoup(json_string.text, "html.parser")
        d = [el['href'] for el in bs.find_all("a", {"title": f"{datee}.xlsx"})]
        if d:
            data = d
            break
    if not data:
        print('рассписание отсутствует')
        return ''

    response = session.get(d[0])
    file_Path = 'Sample.xlsx'

    if response.status_code == 200:
        with open(file_Path, 'wb') as file:
            file.write(response.content)
        with open('date.txt', 'w') as f:
            f.write(str(dt_to))
        print('File downloaded successfully')
    else:
        print('Failed to download file')

    session.close()


def read_excel_with_merged_cells(excel_file, classs):
    workbook = load_workbook(excel_file)
    for el in workbook.sheetnames:
        if classs in el:
            classs = el
            break
    workbook.active = workbook[classs]
    sheet = workbook.active
    table_data = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    iter_row = max_row
    iter_col = max_col

    for row in range(1, iter_row + 1):
        row_data = []
        for col in range(1, iter_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value

            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row = merged_range.min_col, merged_range.min_row
                max_col, max_row = merged_range.max_col, merged_range.max_row

                if min_row <= row <= max_row and min_col <= col <= max_col:
                    top_left_cell = sheet.cell(row=min_row, column=min_col)
                    cell_value = top_left_cell.value
                    break

            row_data.append(cell_value)
        table_data.append(row_data)
    return table_data


def all_day(shool_class):
    emojis_digits = ['0️⃣', '1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣', '6️⃣', '7️⃣', '8️⃣', '9️⃣']
    global classes_A
    global classes_B
    sort_table = read_excel_with_merged_cells("Sample.xlsx", shool_class)
    sort_table = sort_table[1:]

    for i in range(len(sort_table)):
        sort_table[i].pop(0)

    sort_table_all = sort_table

    classes_A = {}
    classes_B = {}
    classes = {shool_class: []}
    classs = []

    for el in sort_table:
        if not el[1]:
            for word in el[1:]:
                if word and not word in classs:
                    classs.append(word)

    for el in sort_table_all[0]:
        if el:
            for char in el:
                if char.isdigit():
                    if el not in classes_A and el not in classes_B:
                        classes_A[el] = None
                        classes_B[el] = None
                        if el in sort_table[0]:
                            classs.append(el)
                        break
    classes[shool_class] = classs
    dop_sort = []

    for cl in list(classes_A.keys()):
        rasp_A = []
        rasp_B = []
        flag = 1

        for el in sort_table_all[2:]:
            if el[0] and flag:
                rasp_A.append(
                    f'{emojis_digits[int(el[0])]} {str(el[1]).replace('\n', '')}\n{el[sort_table_all[0].index(cl)]}')
                rasp_B.append(
                    f'{emojis_digits[int(el[0])]} {str(el[1]).replace('\n', '')}\n{el[sort_table_all[0].index(cl) + 1]}')
            else:
                if flag:
                    for i in range(2, len(el), 2):
                        if el[i]:
                            classes_A[el[i]] = None
                            classes_B[el[i + 1]] = None
                    flag = 0

            if flag == 0 and not el in dop_sort:
                dop_sort.append(el)

            data_A = f'\n----------------------\n{'\n----------------------\n'.join(rasp_A)}'.replace('None',
                                                                                                      '\nПусто').replace(
                '\n\n', '\n')

            classes_A[cl] = data_A

            data_B = f'\n----------------------\n{'\n----------------------\n'.join(rasp_B)}'.replace('None',
                                                                                                      '\nПусто').replace(
                '\n\n', '\n')

            classes_B[cl] = data_B
    print(dop_sort)
    if dop_sort:
        for i in range(1, len(dop_sort[0]), 2):
            dop_A = []
            dop_B = []
            fl = 1

            for obj in dop_sort[1:]:

                if fl:
                    dop_A.append(f'{emojis_digits[int(obj[0])]} {obj[1]} {obj[i]}')
                    dop_B.append(f'{emojis_digits[int(obj[0])]} {obj[1]} {obj[i+1]}')
                    continue
                dop_A.append(str(obj[i]))
                dop_B.append(str(obj[i + 1]))
            print(dop_A)
            if dop_A:
                classes_A[
                    dop_sort[0][i]] = f'\n----------------------\n{'\n----------------------\n'.join(dop_A)}'.replace(
                    'None', '\nПусто').replace('\n\n', '\n')
            if dop_B:
                classes_B[dop_sort[0][
                    i + 1]] = f'\n----------------------\n{'\n----------------------\n'.join(dop_B)}'.replace('None',
                                                                                                              '\nПусто').replace(
                    '\n\n', '\n')


    with open('date.txt', 'r') as f:
        date = f.readline()
    data = {'A': classes_A, 'B': classes_B, 'date': date, 'classes': classes}

    with open(f'data_{shool_class}.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
        with open(f'data_{shool_class}.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

    data_A = {}
    data_B = {}

    for clas in {'10', '11'}:
        with open(f'data_{clas}.json', 'r', encoding='utf-8') as f:
            file = json.load(f)
            data_A = data_A | file['A']
            data_B = data_B | file['B']

    data = {'A': data_A, 'B': data_B}
    with open(f'data_all.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


    




